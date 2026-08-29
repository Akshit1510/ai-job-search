"""Select postings from a job-findings workbook by minimum fit.

Reads the ``Findings`` sheet of a workbook produced by ``export_jobs_xlsx.py``
and returns the rows whose fit is at least as good as a given threshold
(High > Medium > Low). Used by the ``/apply_from_excel`` command to build the
list of postings to run the ``/apply`` workflow against.

Usage:
    python scripts/select_jobs_from_excel.py out/reports/job_findings_2026-08-17.xlsx
    python scripts/select_jobs_from_excel.py job_findings_2026-08-17.xlsx --threshold High
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

from openpyxl import load_workbook

from export_jobs_xlsx import FIT_ORDER

FINDINGS_SHEET = "Findings"
DEFAULT_REPORTS_DIR = Path("out/reports")

# Column positions in the Findings sheet, matching FINDINGS_COLUMNS in export_jobs_xlsx.py.
COL_FIT = 2
COL_TITLE = 3
COL_COMPANY = 4
COL_LOCATION = 5
COL_TIER = 6
COL_TRACK = 8
COL_SENIORITY = 9
COL_NOTES = 10
COL_URL = 11


def resolve_workbook_path(raw: str) -> Path:
    """Resolve a bare filename against out/reports/, appending .xlsx if missing."""
    path = Path(raw)
    if path.suffix.lower() != ".xlsx":
        path = path.with_suffix(".xlsx")
    if not path.is_file() and not path.is_absolute() and len(path.parts) == 1:
        candidate = DEFAULT_REPORTS_DIR / path.name
        if candidate.is_file():
            return candidate
    return path


def load_findings(path: Path) -> list[dict]:
    """Read every row of the Findings sheet into a list of dicts."""
    workbook = load_workbook(path)
    if FINDINGS_SHEET not in workbook.sheetnames:
        raise ValueError(f"workbook has no '{FINDINGS_SHEET}' sheet: {path}")
    sheet = workbook[FINDINGS_SHEET]

    jobs = []
    for row in range(2, sheet.max_row + 1):
        fit = sheet.cell(row=row, column=COL_FIT).value
        title = sheet.cell(row=row, column=COL_TITLE).value
        if not fit or not title:
            continue
        url_cell = sheet.cell(row=row, column=COL_URL)
        url = url_cell.hyperlink.target if url_cell.hyperlink else url_cell.value
        jobs.append(
            {
                "fit": fit,
                "title": title,
                "company": sheet.cell(row=row, column=COL_COMPANY).value or "",
                "location": sheet.cell(row=row, column=COL_LOCATION).value or "",
                "location_tier": sheet.cell(row=row, column=COL_TIER).value or "",
                "track": sheet.cell(row=row, column=COL_TRACK).value or "",
                "seniority": sheet.cell(row=row, column=COL_SENIORITY).value or "",
                "notes": sheet.cell(row=row, column=COL_NOTES).value or "",
                "url": url or "",
            }
        )
    return jobs


def filter_by_threshold(jobs: list[dict], threshold: str) -> list[dict]:
    """Keep jobs whose fit is at least as good as threshold (High > Medium > Low)."""
    threshold_rank = FIT_ORDER.get(threshold.capitalize())
    if threshold_rank is None:
        raise ValueError(f"unknown threshold '{threshold}', expected High/Medium/Low")
    return [j for j in jobs if FIT_ORDER.get(j["fit"], 99) <= threshold_rank]


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("workbook", help="Path to a job-findings .xlsx (bare filename resolves under out/reports/)")
    parser.add_argument(
        "--threshold",
        default="Medium",
        help="Minimum fit to include: High, Medium, or Low (default: Medium)",
    )
    args = parser.parse_args(argv)

    path = resolve_workbook_path(args.workbook)
    if not path.is_file():
        print(f"error: no such workbook: {path}", file=sys.stderr)
        return 1

    try:
        jobs = load_findings(path)
        selected = filter_by_threshold(jobs, args.threshold)
    except ValueError as exc:
        print(f"error: {exc}", file=sys.stderr)
        return 1

    # ensure_ascii=True (the default) keeps output pure-ASCII (\uXXXX escapes for
    # non-ASCII names like "Schrödinger") so it survives stdout redirection under
    # Windows' non-UTF-8 console encoding; json.loads decodes it back correctly.
    print(json.dumps({"workbook": str(path), "threshold": args.threshold.capitalize(), "jobs": selected}))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
