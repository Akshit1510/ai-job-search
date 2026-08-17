"""Tests for the run-report to Excel exporter."""

from __future__ import annotations

import json
from pathlib import Path

import pytest
from openpyxl import load_workbook

from export_jobs_xlsx import export, linkedin_people_search, main, sort_jobs


@pytest.fixture
def report() -> dict:
    return {
        "run_date": "2026-08-11",
        "focus": "Named target companies",
        "portals_run": ["linkedin-search"],
        "portals_skipped": [{"portal": "jobnet-search", "reason": "out of market"}],
        "portal_health": [],
        "jobs": [
            {
                "fit": "Low",
                "title": "Staff Engineer",
                "company": "Zeta",
                "location": "Bengaluru",
                "location_tier": "Tier 2",
                "posted": "2026-07-26",
                "track": "Backend",
                "seniority": "Staff - above band",
                "notes": "Above band.",
                "url": "https://example.com/3",
            },
            {
                "fit": "High",
                "title": "Software Engineer II",
                "company": "Rippling",
                "location": "Bengaluru",
                "location_tier": "Tier 2",
                "posted": "2026-07-31",
                "track": "Both",
                "seniority": "exact band",
                "notes": "Two-track hit.",
                "url": "https://example.com/1",
            },
            {
                "fit": "Medium",
                "title": "SWE III",
                "company": "Alpha",
                "location": "Pune",
                "location_tier": "Tier 1",
                "posted": "2026-08-04",
                "track": "AI",
                "seniority": "in band",
                "notes": "Partial match.",
                "url": "https://example.com/2",
            },
        ],
        "observations": ["Backend weighting change surfaced two new matches."],
    }


def test_sort_jobs_orders_high_then_medium_then_low(report):
    assert [job["fit"] for job in sort_jobs(report["jobs"])] == ["High", "Medium", "Low"]


def test_sort_jobs_is_stable_across_equal_fit():
    jobs = [
        {"fit": "High", "company": "Zeta"},
        {"fit": "High", "company": "Alpha"},
    ]
    assert [job["company"] for job in sort_jobs(jobs)] == ["Alpha", "Zeta"]


def test_linkedin_people_search_url_encodes_spaces_and_symbols():
    url = linkedin_people_search("Rippling AI & Platform")
    assert "Rippling%20AI%20%26%20Platform" in url
    assert url.startswith("https://www.linkedin.com/search/results/people/")


def test_export_creates_three_sheets(report, tmp_path):
    out = export(report, tmp_path / "findings.xlsx")
    book = load_workbook(out)
    assert book.sheetnames == ["Findings", "Highlights", "Run Info"]


def test_findings_sheet_has_one_row_per_job_sorted_by_fit(report, tmp_path):
    out = export(report, tmp_path / "findings.xlsx")
    sheet = load_workbook(out)["Findings"]
    assert sheet.max_row == len(report["jobs"]) + 1  # + header
    assert [sheet.cell(row=r, column=2).value for r in range(2, 5)] == [
        "High",
        "Medium",
        "Low",
    ]


def test_posting_url_is_a_live_hyperlink(report, tmp_path):
    out = export(report, tmp_path / "findings.xlsx")
    sheet = load_workbook(out)["Findings"]
    assert sheet.cell(row=2, column=11).hyperlink.target == "https://example.com/1"


def test_low_fit_rows_omit_referral_links(report, tmp_path):
    """The /scrape skill limits referral links to High and Medium fit."""
    out = export(report, tmp_path / "findings.xlsx")
    sheet = load_workbook(out)["Findings"]
    assert sheet.cell(row=2, column=12).value == "Recruiters"  # High
    assert sheet.cell(row=4, column=12).value is None  # Low


def test_highlights_sheet_contains_only_high_fit(report, tmp_path):
    out = export(report, tmp_path / "findings.xlsx")
    sheet = load_workbook(out)["Highlights"]
    assert sheet.max_row == 2  # header + the single High job
    assert sheet.cell(row=2, column=1).value == "Rippling"


def test_run_info_reports_fit_counts_and_skipped_portals(report, tmp_path):
    out = export(report, tmp_path / "findings.xlsx")
    sheet = load_workbook(out)["Run Info"]
    values = {sheet.cell(row=r, column=1).value: sheet.cell(row=r, column=2).value
              for r in range(2, sheet.max_row + 1)}
    assert values["High / Medium / Low"] == "1 / 1 / 1"
    assert values["Skipped: jobnet-search"] == "out of market"
    assert values["Total postings"] == "3"


def test_main_writes_file_named_from_run_date(report, tmp_path):
    src = tmp_path / "report.json"
    src.write_text(json.dumps(report), encoding="utf-8")
    assert main([str(src), "--out-dir", str(tmp_path / "out")]) == 0
    assert (tmp_path / "out" / "job_findings_2026-08-11.xlsx").is_file()


def test_main_returns_error_for_missing_report(tmp_path, capsys):
    assert main([str(tmp_path / "nope.json")]) == 1
    assert "no such report" in capsys.readouterr().err


def test_export_creates_missing_output_directory(report, tmp_path):
    out = export(report, tmp_path / "deep" / "nested" / "f.xlsx")
    assert out.is_file()


def test_real_run_report_exports_cleanly(tmp_path):
    """Guards the checked-in report against schema drift."""
    path = Path(__file__).resolve().parent.parent / "job_scraper" / "run_reports" / "2026-08-11.json"
    if not path.is_file():
        pytest.skip("run report not present")
    data = json.loads(path.read_text(encoding="utf-8"))
    sheet = load_workbook(export(data, tmp_path / "real.xlsx"))["Findings"]
    assert sheet.max_row == len(data["jobs"]) + 1
