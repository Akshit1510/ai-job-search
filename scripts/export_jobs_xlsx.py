"""Convert a job-scraper run report (JSON) into a formatted Excel workbook.

A run report is produced by `/scrape` and stored under
``job_scraper/run_reports/<YYYY-MM-DD>.json``. This script turns one into a
three-sheet workbook under ``out/reports/``:

    Findings   - one row per posting, sorted High > Medium > Low
    Highlights - the analyst notes for every High-fit posting
    Run Info   - portals used/skipped and cross-run observations

Usage:
    python scripts/export_jobs_xlsx.py job_scraper/run_reports/2026-08-11.json
    python scripts/export_jobs_xlsx.py <report.json> --out-dir out/reports
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path
from urllib.parse import quote

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# Fit bands, most promising first. Drives both sort order and row colour.
FIT_ORDER = {"High": 0, "Medium": 1, "Low": 2}

FIT_FILL = {
    "High": PatternFill("solid", fgColor="D6EFD8"),
    "Medium": PatternFill("solid", fgColor="FDF3D2"),
    "Low": PatternFill("solid", fgColor="F2F2F2"),
}

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)

FINDINGS_COLUMNS = [
    ("#", 5),
    ("Fit", 9),
    ("Title", 42),
    ("Company", 13),
    ("Location", 22),
    ("Tier", 8),
    ("Posted", 12),
    ("Track", 10),
    ("Seniority vs Band", 30),
    ("Notes", 80),
    ("Posting URL", 46),
    ("Recruiter Search", 46),
    ("Peer Search", 46),
]


def linkedin_people_search(keywords: str) -> str:
    """Build a LinkedIn people-search URL for referral outreach.

    Link generation only - the workbook never contains looked-up contacts.
    """
    return (
        "https://www.linkedin.com/search/results/people/?keywords="
        f"{quote(keywords)}&origin=GLOBAL_SEARCH_HEADER"
    )


def _style_header(sheet: Worksheet, columns: list[tuple[str, int]]) -> None:
    for index, (title, width) in enumerate(columns, start=1):
        cell = sheet.cell(row=1, column=index, value=title)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", horizontal="center")
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.freeze_panes = "A2"


def _hyperlink(sheet: Worksheet, row: int, column: int, url: str, label: str) -> None:
    cell = sheet.cell(row=row, column=column, value=label)
    cell.hyperlink = url
    cell.font = Font(color="0563C1", underline="single")


def sort_jobs(jobs: list[dict]) -> list[dict]:
    """High-fit first, then by company so a company's roles stay adjacent."""
    return sorted(
        jobs,
        key=lambda job: (
            FIT_ORDER.get(job.get("fit", "Low"), 99),
            job.get("company", ""),
        ),
    )


def build_findings_sheet(sheet: Worksheet, jobs: list[dict]) -> None:
    sheet.title = "Findings"
    _style_header(sheet, FINDINGS_COLUMNS)

    for row, job in enumerate(sort_jobs(jobs), start=2):
        fit = job.get("fit", "Low")
        company = job.get("company", "")
        title = job.get("title", "")

        values = [
            row - 1,
            fit,
            title,
            company,
            job.get("location", ""),
            job.get("location_tier", ""),
            job.get("posted", ""),
            job.get("track", ""),
            job.get("seniority", ""),
            job.get("notes", ""),
        ]
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row=row, column=column, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=(column == 10))
            if fit in FIT_FILL:
                cell.fill = FIT_FILL[fit]

        _hyperlink(sheet, row, 11, job.get("url", ""), "Open posting")
        # Referral links only for High/Medium, matching the /scrape skill rule.
        if fit in ("High", "Medium"):
            _hyperlink(
                sheet, row, 12, linkedin_people_search(f"{company} recruiter"), "Recruiters"
            )
            _hyperlink(
                sheet, row, 13, linkedin_people_search(f"{company} {title}"), "Team peers"
            )
        sheet.row_dimensions[row].height = 58

    sheet.auto_filter.ref = f"A1:{get_column_letter(len(FINDINGS_COLUMNS))}{len(jobs) + 1}"


def build_highlights_sheet(sheet: Worksheet, jobs: list[dict]) -> None:
    columns = [("Company", 14), ("Title", 44), ("Location", 24), ("Why it matters", 110)]
    _style_header(sheet, columns)

    row = 2
    for job in sort_jobs(jobs):
        if job.get("fit") != "High":
            continue
        for column, value in enumerate(
            [
                job.get("company", ""),
                job.get("title", ""),
                job.get("location", ""),
                job.get("notes", ""),
            ],
            start=1,
        ):
            cell = sheet.cell(row=row, column=column, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=(column == 4))
        sheet.row_dimensions[row].height = 92
        row += 1


def build_run_info_sheet(sheet: Worksheet, report: dict) -> None:
    columns = [("Field", 26), ("Value", 120)]
    _style_header(sheet, columns)

    jobs = report.get("jobs", [])
    counts = {band: sum(1 for j in jobs if j.get("fit") == band) for band in FIT_ORDER}

    rows: list[tuple[str, str]] = [
        ("Run date", report.get("run_date", "")),
        ("Focus", report.get("focus", "")),
        ("Total postings", str(len(jobs))),
        ("High / Medium / Low", f"{counts['High']} / {counts['Medium']} / {counts['Low']}"),
        ("Portals run", ", ".join(report.get("portals_run", []))),
    ]
    for skipped in report.get("portals_skipped", []):
        rows.append((f"Skipped: {skipped['portal']}", skipped.get("reason", "")))
    for health in report.get("portal_health", []):
        rows.append((f"Health: {health.get('portal', '')}", health.get("status", "")))
    for index, observation in enumerate(report.get("observations", []), start=1):
        rows.append((f"Observation {index}", observation))

    for row, (field, value) in enumerate(rows, start=2):
        sheet.cell(row=row, column=1, value=field).font = Font(bold=True)
        cell = sheet.cell(row=row, column=2, value=value)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        sheet.row_dimensions[row].height = 34


def export(report: dict, out_path: Path) -> Path:
    workbook = Workbook()
    build_findings_sheet(workbook.active, report.get("jobs", []))
    build_highlights_sheet(workbook.create_sheet("Highlights"), report.get("jobs", []))
    build_run_info_sheet(workbook.create_sheet("Run Info"), report)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(out_path)
    return out_path


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("report", type=Path, help="Path to a run-report JSON file")
    parser.add_argument(
        "--out-dir", type=Path, default=Path("out/reports"), help="Output directory"
    )
    args = parser.parse_args(argv)

    if not args.report.is_file():
        print(f"error: no such report: {args.report}", file=sys.stderr)
        return 1

    report = json.loads(args.report.read_text(encoding="utf-8"))
    out_path = args.out_dir / f"job_findings_{report.get('run_date', args.report.stem)}.xlsx"
    export(report, out_path)
    print(f"wrote {out_path} ({len(report.get('jobs', []))} postings)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
